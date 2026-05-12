import json,re,urllib.request
from pathlib import Path
from datetime import datetime,date
from openpyxl import load_workbook

SHEET_ID='1bG6RsGZK39lNogtTBS5_voJz0WVcRu2D'
URL=f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx'
TMP=Path('luz-teddy/STATE/source.xlsx')
OUT=Path('luz-teddy/data/may-june-events.json')
MONTHS={5,6}
DATE_RE=re.compile(r'(\d{1,2})[./](\d{1,2})(?:[./](\d{2,4}))?')
TIME_RE=re.compile(r'(\d{1,2}:\d{2})(?:\s*[-–]\s*(\d{1,2}:\d{2}))?')
SPLIT_RE=re.compile(r'\n|\r|•|▪|▫|;|\s{3,}')
SKIP=('לוח','גלישת טקסט','גודל גופן','כל חודש')

def clean(v):
    if v is None: return ''
    if isinstance(v,(datetime,date)): return v.strftime('%d.%m.%Y')
    return re.sub(r'\s+',' ',str(v).strip())

def kind(s):
    x=s.lower()
    if 'מבחן' in x: return 'exam'
    if 'בוחן' in x or 'בחן' in x: return 'quiz'
    if 'חופש' in x or 'חופשה' in x: return 'vacation'
    if 'טיול' in x or 'סיור' in x: return 'trip'
    if 'חוג' in x: return 'club'
    if 'טקס' in x: return 'ceremony'
    if 'מפגש' in x or 'ישיבה' in x: return 'meeting'
    if 'הגשה' in x or 'יעד' in x: return 'deadline'
    return 'activity'

def parse_date(s):
    if isinstance(s,(datetime,date)): return s.date() if isinstance(s,datetime) else s
    m=DATE_RE.search(clean(s))
    if not m: return None
    d,mo,y=m.groups(); d=int(d); mo=int(mo)
    if y: y=int(y); y=2000+y if y<100 else y
    else:
        now=datetime.now(); base=now.year if now.month>=9 else now.year-1; y=base if mo>=9 else base+1
    try: return date(y,mo,d)
    except ValueError: return None

def parse_time(s):
    m=TIME_RE.search(clean(s))
    if not m: return ('','')
    return (m.group(1),m.group(2) or '')

def bad(s):
    x=clean(s)
    return (not x) or len(x)<2 or any(k in x for k in SKIP)

def title_of(s):
    x=TIME_RE.sub('',clean(s))
    x=re.sub(r'^(מפגש|ישיבה|פעילות|אירוע)\s*[:\-–]?\s*','',x).strip()
    return x or clean(s)

def add_event(out,text,dt,source):
    if not dt or dt.month not in MONTHS or bad(text): return
    start,end=parse_time(text)
    title=title_of(text)
    if bad(title): return
    out.append({'date':dt.isoformat(),'title':title,'startTime':start,'endTime':end,'type':kind(text),'source':source})

def scan_sheet(ws):
    out=[]; col_date={}
    for row in ws.iter_rows(values_only=True):
        vals=[clean(v) for v in row]
        for c,v in enumerate(vals):
            if not v: continue
            dt=parse_date(v)
            if dt:
                col_date[c]=dt
                rest=DATE_RE.sub('',v)
                rest=re.sub(r'^יום\s+\S+','',rest).strip()
                add_event(out,rest,dt,ws.title)
            elif c in col_date and not bad(v):
                for part in SPLIT_RE.split(v): add_event(out,part,col_date[c],ws.title)
    return out

def main():
    TMP.parent.mkdir(parents=True,exist_ok=True); OUT.parent.mkdir(parents=True,exist_ok=True)
    req=urllib.request.Request(URL,headers={'User-Agent':'luz-teddy-extractor'})
    data=urllib.request.urlopen(req,timeout=60).read()
    TMP.write_bytes(data)
    wb=load_workbook(TMP,data_only=True,read_only=True)
    events=[]
    for ws in wb.worksheets: events.extend(scan_sheet(ws))
    seen=set(); uniq=[]
    for e in sorted(events,key=lambda x:(x['date'],x.get('startTime',''),x['title'])):
        k=(e['date'],e['startTime'],e['endTime'],e['title'])
        if k not in seen: seen.add(k); uniq.append(e)
    OUT.write_text(json.dumps({'sourceSheetId':SHEET_ID,'months':[5,6],'count':len(uniq),'events':uniq},ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print('MAY_JUNE_EVENTS',len(uniq))

if __name__=='__main__': main()
